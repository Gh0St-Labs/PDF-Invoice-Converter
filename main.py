import pandas as pd
import glob
import openpyxl
from fpdf import FPDF
from pathlib import Path

files = glob.glob('Invoice\\*.xlsx')


for file in files:
    data = pd.read_excel(file, sheet_name='Sheet 1')
    pdf = FPDF(orientation="P", unit="mm",format="A4")
    pdf.set_auto_page_break(auto=False, margin=0)
    pdf.add_page()
    extract_filename = Path(file).stem
    invoice_no = extract_filename.split("-")[0]
    date = extract_filename.split("-")[1]
    pdf.set_font(family="Times", style="B", size=17)
    pdf.cell(w=50, h=18, txt=f"Invoice No: {invoice_no}", ln=1)
    pdf.cell(w=70, h=18, txt=f"Date: {date}", ln=1)
    pdf.set_font(family="Times", size=17)

    columns = data.columns
    pdf.set_font(family="Times", size=12)
    pdf.set_text_color(0, 0, 0)
    columns = [item.replace("_"," ").title() for item in columns]
    pdf.cell(w=30, h=15, align="C", txt=columns[0], border=1)
    pdf.cell(w=75, h=15, align="C", txt=columns[1], border=1)
    pdf.cell(w=35, h=15, align="C", txt=columns[2], border=1)
    pdf.cell(w=30, h=15, align="C", txt=columns[3], border=1)
    pdf.cell(w=20, h=15, align="C", txt=columns[4], border=1, ln=1)

    for index, rows in data.iterrows():
        pdf.set_font(family="Times", size=12)
        pdf.set_text_color(80,80,80)
        pdf.cell(w=30, h=15, align="C", txt=str(rows['product_id']), border=1)
        pdf.cell(w=75, h=15, align="C", txt=str(rows['product_name']), border=1)
        pdf.cell(w=35, h=15, align="C", txt=str(rows['amount_purchased']), border=1)
        pdf.cell(w=30, h=15, align="C", txt=str(rows['price_per_unit']), border=1)
        pdf.cell(w=20, h=15, align="C", txt=str(rows['total_price']), border=1, ln=1)

    total = data['total_price'].sum()
    pdf.cell(w=0, h=15, align="C", txt=f"Total is {str(total)}", border=1, ln=1)

    pdf.set_font(family="Times", style="I", size=13)
    pdf.set_text_color(0,0,0)
    pdf.write(h=14, txt="Python")
    pdf.image(w=5, h=5, x=25, y=111, name="kisspng-python-computer-icons-tutorial-computer-programmin-social-icons-5ad5ccbb30c4a8.2707803315239610191998.jpg")


    pdf.output(f"Converted_PDFs\\{extract_filename}.pdf")

